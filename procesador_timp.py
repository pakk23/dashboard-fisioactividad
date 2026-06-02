"""
Procesador de exportaciones TIMP para Fisioactividad.
Versión sincronizada con skill informe-reservas-y-facturacion v2.
"""
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Clasificación de servicios ───────────────────────────────────────────────
SERVICIOS_INDIVIDUAL = ["fisioterapia", "pilates individual", "fisioactividad"]
SERVICIOS_DUO        = ["dúo pilates", "duo pilates", "dúo fisioactividad", "duo fisioactividad"]
SERVICIO_FISIO_ECO   = "fisioterapia manual, invasiva y ecográfica"
SERVICIO_REEVALUACION = "reevaluación"
SERVICIO_ONCOLOGICO  = "aecc"

SERVICIOS_EXCLUIR = [
    "cita exterior", "gestión", "gestion",
    "formación interna", "formacion interna",
    "reunión cliente", "reunion cliente",
]

# Capacidades por servicio para hoja Sesiones
CAPACIDADES = {
    "pilates studio": 5,
    "pilates terapéutico": 5,
    "pilates terapeutico": 5,
    "ejercicio terapéutico": 6,
    "ejercicio terapeutico": 6,
    "acondicionamiento físico": 6,
    "acondicionamiento fisico": 6,
    "dúo pilates": 2,
    "duo pilates": 2,
    "dúo fisioactividad": 2,
    "duo fisioactividad": 2,
}
CAP_TALLER   = 15
CAP_DEFAULT  = 1

MESES_ES = {1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",
            7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}
DIAS_ES  = {0:"lunes",1:"martes",2:"miércoles",3:"jueves",4:"viernes",5:"sábado",6:"domingo"}

# Tarifas de referencia 2026
TARIFAS = {
    "Grupal mensualidad": {"venta": "Mensualidad", "precio": 150,  "sesiones": 10, "precio_reserva": 15.0},
    "Grupal bono 20":     {"venta": "Bono 20",     "precio": 420,  "sesiones": 20, "precio_reserva": 21.0},
    "Dúo bono 20":        {"venta": "Bono 20",     "precio": 600,  "sesiones": 20, "precio_reserva": 30.0},
    "Dúo bono 40":        {"venta": "Bono 40",     "precio": 1000, "sesiones": 40, "precio_reserva": 25.0},
    "Individual sesión":  {"venta": "Individual",  "precio": 60,   "sesiones": 1,  "precio_reserva": 60.0},
    "Individual bono 10": {"venta": "Bono 10",     "precio": 530,  "sesiones": 10, "precio_reserva": 53.0},
    "Individual bono 20": {"venta": "Bono 20",     "precio": 1000, "sesiones": 20, "precio_reserva": 50.0},
}
TARIFA_DEFAULT_GRUPAL = TARIFAS["Grupal mensualidad"]
TARIFA_DEFAULT_DUO    = TARIFAS["Dúo bono 20"]
TARIFA_DEFAULT_IND    = TARIFAS["Individual sesión"]


def _norm(s):
    if not isinstance(s, str):
        return ""
    return (s.lower()
            .replace("á","a").replace("é","e").replace("í","i")
            .replace("ó","o").replace("ú","u").replace("ü","u")
            .strip())


def _tipo_servicio(servicio):
    s = _norm(servicio)
    if any(t in s for t in SERVICIOS_DUO):
        return "duo"
    if any(t in s for t in SERVICIOS_INDIVIDUAL):
        return "individual"
    return "grupal"


def _capacidad(servicio):
    s = _norm(servicio)
    if "taller" in s:
        return CAP_TALLER
    for key, cap in CAPACIDADES.items():
        if key in s:
            return cap
    return CAP_DEFAULT


# ── Lectura ───────────────────────────────────────────────────────────────────
def leer_archivo(file) -> pd.DataFrame:
    raw = pd.read_excel(file, header=None)
    header_row = 0
    for i in range(min(10, len(raw))):
        if str(raw.iloc[i, 0]).strip() == "ID":
            header_row = i
            break
    df = pd.read_excel(file, header=header_row)
    df.columns = df.columns.str.strip()
    df = df.dropna(subset=["ID"])
    for col in ["Inicio", "Fin", "Fecha de reserva", "Fecha cancelación"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    for col in ["Precio", "Precio reserva", "Sesiones del bono"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["_editada"] = False
    return df


# ── Paso 0: Reestructurar columnas ───────────────────────────────────────────
def paso0_reestructurar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    # 0a. Columna "Cliente" = "Apellidos, Nombre"
    nombre_col    = "Cliente (nombre)"
    apellidos_col = "Cliente (apellidos)"
    if nombre_col in df.columns and apellidos_col in df.columns:
        df["Cliente"] = (
            df[apellidos_col].fillna("").str.strip() + ", " +
            df[nombre_col].fillna("").str.strip()
        ).str.strip(", ")
        # Insertar "Cliente" donde estaba "Cliente (nombre)"
        pos = df.columns.get_loc(nombre_col)
        cols = list(df.columns)
        cols.remove("Cliente")
        cols.insert(pos, "Cliente")
        df = df[cols]
        df = df.drop(columns=[nombre_col, apellidos_col])

    # 0b. Separar "Inicio" en "Fecha" (texto DD/MM/YYYY) y "Hora" (HH:MM)
    if "Inicio" in df.columns:
        df["Fecha"] = df["Inicio"].dt.strftime("%d/%m/%Y")
        df["Hora"]  = df["Inicio"].dt.strftime("%H:%M")
        # Reemplazar columna "Fin" con la hora, renombrarla a "Hora"
        if "Fin" in df.columns:
            pos_fin = df.columns.get_loc("Fin")
            df = df.drop(columns=["Fin"])
            cols = list(df.columns)
            cols.remove("Hora")
            cols.insert(pos_fin, "Hora")
            df = df[cols]
        # Reemplazar "Inicio" con "Fecha"
        pos_ini = df.columns.get_loc("Inicio")
        df = df.drop(columns=["Inicio"])
        cols = list(df.columns)
        cols.remove("Fecha")
        cols.insert(pos_ini, "Fecha")
        df = df[cols]
    return df


# ── Paso 1: Separar a hojas ───────────────────────────────────────────────────
def paso1_separar_hojas(df: pd.DataFrame):
    """
    Extrae filas a hojas separadas y las elimina de Sheet1.
    Devuelve (df_main, df_fisio_eco, df_excluidas, df_canceladas).
    """
    servicio_norm = df["Servicio"].fillna("").apply(_norm)
    estado        = df["Estado de reserva"].fillna("")
    canjeada      = df["Canjeada"].fillna("")

    # Fisio Invasiva
    mask_eco = servicio_norm.str.contains(SERVICIO_FISIO_ECO, na=False)
    df_eco   = df[mask_eco].copy()
    df_eco["Ingreso estimado"] = 60.0

    # Excluidas
    mask_excl = servicio_norm.isin(SERVICIOS_EXCLUIR)
    df_excl   = df[mask_excl & ~mask_eco].copy()

    # Canceladas / En cola
    mask_cancel = (
        ((estado == "Cancelada") & (canjeada.isin(["No", "Regalada"]))) |
        (estado == "En cola")
    )
    df_cancel = df[mask_cancel & ~mask_eco & ~mask_excl].copy()

    # Sheet1 limpio
    mask_keep = ~(mask_eco | mask_excl | mask_cancel)
    df_main   = df[mask_keep].copy()

    return df_main, df_eco, df_excl, df_cancel


# ── Paso 2: Oncológico AECC ───────────────────────────────────────────────────
def paso2_oncologico(df: pd.DataFrame) -> pd.DataFrame:
    mask_aecc = (
        df["Servicio"].fillna("").apply(_norm).str.contains(SERVICIO_ONCOLOGICO, na=False) |
        df["Venta"].fillna("").apply(_norm).str.contains(SERVICIO_ONCOLOGICO, na=False)
    )
    if not mask_aecc.any():
        return df

    mask_no_asistido = mask_aecc & (df["Pasar lista"].fillna("") == "No asistido")
    df = df[~mask_no_asistido].copy()

    mask_aecc = (
        df["Servicio"].fillna("").apply(_norm).str.contains(SERVICIO_ONCOLOGICO, na=False) |
        df["Venta"].fillna("").apply(_norm).str.contains(SERVICIO_ONCOLOGICO, na=False)
    )
    if not mask_aecc.any():
        return df

    fecha_col = "Fecha" if "Fecha" in df.columns else "Inicio"
    sesiones_unicas = df.loc[mask_aecc, fecha_col].nunique()
    total_reservas  = mask_aecc.sum()
    precio_total    = sesiones_unicas * 45.0
    precio_reserva  = round(precio_total / total_reservas, 2) if total_reservas > 0 else 0.0

    df.loc[mask_aecc, "Venta"]           = "AECC"
    df.loc[mask_aecc, "Precio"]          = precio_total
    df.loc[mask_aecc, "Sesiones del bono"] = total_reservas
    df.loc[mask_aecc, "Precio reserva"]  = precio_reserva
    df.loc[mask_aecc, "_editada"]        = True
    return df


# ── Paso 3: Reevaluación ──────────────────────────────────────────────────────
def paso3_reevaluacion(df: pd.DataFrame) -> pd.DataFrame:
    mask = (
        df["Servicio"].fillna("").apply(_norm).str.contains(SERVICIO_REEVALUACION, na=False) &
        (df["Estado de reserva"] == "Aceptada")
    )
    if not mask.any():
        return df
    df.loc[mask, "Venta"]             = "Individual"
    df.loc[mask, "Precio"]            = 0.0
    df.loc[mask, "Sesiones del bono"] = 1
    df.loc[mask, "Precio reserva"]    = 0.0
    df.loc[mask, "_editada"]          = True
    return df


# ── Paso 4: Rellenar no canjeadas ─────────────────────────────────────────────
def _buscar_tarifa(df: pd.DataFrame, cliente_id, servicio, tipo) -> dict | None:
    canjeadas = df[
        (df["Canjeada"] != "No") &
        (df["Estado de reserva"] == "Aceptada") &
        df["Venta"].notna() & df["Precio"].notna()
    ]
    same = canjeadas[
        (canjeadas["Cliente (ID)"] == cliente_id) &
        (canjeadas["Servicio"] == servicio)
    ]
    if not same.empty:
        r = same.iloc[-1]
        return {"venta": r["Venta"], "precio": r["Precio"],
                "sesiones": r["Sesiones del bono"], "precio_reserva": r["Precio reserva"]}
    same_tipo = canjeadas[canjeadas["Cliente (ID)"] == cliente_id]
    same_tipo = same_tipo[same_tipo["Servicio"].fillna("").apply(_tipo_servicio) == tipo]
    if not same_tipo.empty:
        r = same_tipo.iloc[-1]
        return {"venta": r["Venta"], "precio": r["Precio"],
                "sesiones": r["Sesiones del bono"], "precio_reserva": r["Precio reserva"]}
    tipo_rows = canjeadas[canjeadas["Servicio"].fillna("").apply(_tipo_servicio) == tipo]
    if not tipo_rows.empty:
        venta_comun = tipo_rows["Venta"].mode()
        if not venta_comun.empty:
            muestra = tipo_rows[tipo_rows["Venta"] == venta_comun.iloc[0]].iloc[-1]
            return {"venta": muestra["Venta"], "precio": muestra["Precio"],
                    "sesiones": muestra["Sesiones del bono"], "precio_reserva": muestra["Precio reserva"]}
    if tipo == "individual": return TARIFA_DEFAULT_IND
    if tipo == "duo":        return TARIFA_DEFAULT_DUO
    return TARIFA_DEFAULT_GRUPAL


def paso4_rellenar_no_canjeadas(df: pd.DataFrame) -> pd.DataFrame:
    mask = (df["Canjeada"].fillna("") == "No") & (df["Estado de reserva"] == "Aceptada")
    if not mask.any():
        return df
    for idx in df[mask].index:
        row  = df.loc[idx]
        tipo = _tipo_servicio(row["Servicio"])
        t    = _buscar_tarifa(df, row["Cliente (ID)"], row["Servicio"], tipo)
        if t:
            df.loc[idx, "Venta"]             = t["venta"]
            df.loc[idx, "Precio"]            = t["precio"]
            df.loc[idx, "Sesiones del bono"] = t["sesiones"]
            df.loc[idx, "Precio reserva"]    = t["precio_reserva"]
            df.loc[idx, "_editada"]          = True
    return df


# ── Paso 6: Ventas Pendientes ─────────────────────────────────────────────────
def paso6_ventas_pendientes(df: pd.DataFrame) -> pd.DataFrame:
    mask = (df["Canjeada"].fillna("") == "No") & (df["Estado de reserva"] == "Aceptada")
    pendientes = df[mask].copy()
    if pendientes.empty:
        return pd.DataFrame()

    pendientes = pendientes.sort_values("Fecha" if "Fecha" in pendientes.columns else "ID")
    primera = pendientes.groupby("Cliente (ID)")["Fecha"].min().rename("Fecha primer día en rojo")
    pendientes = pendientes.merge(primera, on="Cliente (ID)", how="left")

    cols = [c for c in [
        "Cliente (ID)", "Cliente (código)", "Cliente",
        "Servicio", "Fecha primer día en rojo", "Venta",
        "Precio", "Sesiones del bono", "Precio reserva",
    ] if c in pendientes.columns]
    resultado = pendientes.drop_duplicates(subset=["Cliente (ID)"])[cols]
    sort_col = "Cliente" if "Cliente" in resultado.columns else cols[0]
    return resultado.sort_values(sort_col).reset_index(drop=True)


# ── Paso 8: Hoja Sesiones ─────────────────────────────────────────────────────
def paso8_sesiones(df: pd.DataFrame) -> pd.DataFrame:
    if "Fecha" not in df.columns or "Hora" not in df.columns:
        return pd.DataFrame()

    df2 = df.copy()
    df2["_fecha_dt"] = pd.to_datetime(df2["Fecha"], format="%d/%m/%Y", errors="coerce")

    group_cols = ["Servicio", "Fecha", "Hora", "Profesional"]
    group_cols = [c for c in group_cols if c in df2.columns]

    agg = df2.groupby(group_cols, dropna=False).agg(
        Reservas=("ID", "count"),
        euro_pers=("Precio reserva", "mean"),
        euro_ses=("Precio reserva", "sum"),
        _fecha_dt=("_fecha_dt", "first"),
    ).reset_index()

    agg["Capacidad"] = agg["Servicio"].apply(_capacidad)
    agg["Ocupación"] = (agg["Reservas"] / agg["Capacidad"]).clip(upper=1.0)
    agg["€/pers."]   = agg["euro_pers"].round(2)
    agg["€/ses."]    = agg["euro_ses"].round(2)

    # Año, mes, día de la semana
    agg["Año"] = agg["_fecha_dt"].dt.year
    agg["Mes"] = agg["_fecha_dt"].dt.month.map(MESES_ES)
    agg["Día"] = agg["_fecha_dt"].dt.dayofweek.map(DIAS_ES)

    agg = agg.sort_values(["_fecha_dt", "Hora"]).reset_index(drop=True)
    agg["Nº"] = agg.index + 1

    cols_out = ["Nº", "Servicio", "Año", "Mes", "Día", "Fecha", "Hora",
                "Profesional", "Capacidad", "Reservas", "Ocupación", "€/pers.", "€/ses."]
    cols_out = [c for c in cols_out if c in agg.columns]
    # Renombrar Servicio → Actividad y Hora → Sesión
    agg = agg.rename(columns={"Servicio": "Actividad", "Hora": "Sesión"})
    cols_out = ["Nº", "Actividad", "Año", "Mes", "Día", "Fecha", "Sesión",
                "Profesional", "Capacidad", "Reservas", "Ocupación", "€/pers.", "€/ses."]
    cols_out = [c for c in cols_out if c in agg.columns]
    return agg[cols_out].reset_index(drop=True)


# ── Pipeline principal ────────────────────────────────────────────────────────
def procesar(file) -> dict:
    df = leer_archivo(file)
    total_original = len(df)

    df = paso0_reestructurar_columnas(df)
    df, df_eco, df_excl, df_cancel = paso1_separar_hojas(df)

    df = paso2_oncologico(df)
    df = paso3_reevaluacion(df)
    df = paso4_rellenar_no_canjeadas(df)

    df_pendientes = paso6_ventas_pendientes(df)
    df_sesiones   = paso8_sesiones(df)

    editadas = int(df["_editada"].sum())

    stats = {
        "total_original":  total_original,
        "eco_separadas":   len(df_eco),
        "excluidas":       len(df_excl),
        "canceladas":      len(df_cancel),
        "editadas":        editadas,
        "pendientes":      len(df_pendientes),
        "total_procesadas": len(df),
    }

    return {
        "df_main":       df,
        "df_eco":        df_eco,
        "df_excl":       df_excl,
        "df_cancel":     df_cancel,
        "df_pendientes": df_pendientes,
        "df_sesiones":   df_sesiones,
        "stats":         stats,
    }


# ── Exportar Excel ────────────────────────────────────────────────────────────
def exportar_excel(resultado: dict) -> bytes:
    df_main      = resultado["df_main"].copy()
    df_eco       = resultado["df_eco"].copy()
    df_excl      = resultado["df_excl"].copy()
    df_cancel    = resultado["df_cancel"].copy()
    df_pendientes = resultado["df_pendientes"].copy()
    df_sesiones  = resultado["df_sesiones"].copy()

    cols_main = [c for c in df_main.columns if c != "_editada"]
    cols_eco  = [c for c in df_eco.columns  if c != "_editada"]
    cols_excl = [c for c in df_excl.columns if c != "_editada"]
    cols_canc = [c for c in df_cancel.columns if c != "_editada"]

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_main[cols_main].to_excel(writer, sheet_name="Sheet1", index=False)

        if not df_eco.empty:
            df_eco_out = df_eco[cols_eco].copy()
            total_row  = {c: "" for c in cols_eco}
            total_row["Cliente"] = "TOTAL"
            total_row["Ingreso estimado"] = df_eco_out["Ingreso estimado"].sum()
            df_eco_out = pd.concat([df_eco_out, pd.DataFrame([total_row])], ignore_index=True)
            df_eco_out.to_excel(writer, sheet_name="Fisio Invasiva", index=False)

        if not df_excl.empty:
            df_excl[cols_excl].to_excel(writer, sheet_name="Excluidas", index=False)

        if not df_cancel.empty:
            df_cancel[cols_canc].to_excel(writer, sheet_name="Canceladas", index=False)

        if not df_pendientes.empty:
            df_pendientes.to_excel(writer, sheet_name="Ventas Pendientes", index=False)

        # Tabla dinámica
        pivot = (
            df_main.groupby("Profesional")
            .agg(Reservas=("ID", "count"), Facturación=("Precio reserva", "sum"))
            .reset_index()
            .sort_values("Facturación", ascending=False)
        )
        pivot["Facturación"] = pivot["Facturación"].round(2)
        pivot.to_excel(writer, sheet_name="Tabla Dinámica", index=False)

        if not df_sesiones.empty:
            df_sesiones.to_excel(writer, sheet_name="Sesiones", index=False)

    # Aplicar estilos
    wb = openpyxl.load_workbook(output)
    _aplicar_estilos(wb, df_main)
    out2 = BytesIO()
    wb.save(out2)
    return out2.getvalue()


def _aplicar_estilos(wb, df_main):
    ROJO_FONDO   = PatternFill("solid", fgColor="FFD9D9")
    ROJO_TEXTO   = Font(color="CC0000")
    CAB_ROJA     = PatternFill("solid", fgColor="CC0000")
    CAB_OSCURA   = PatternFill("solid", fgColor="2C3E50")
    TEXTO_BLANCO = Font(color="FFFFFF", bold=True)
    FMT_EURO     = '#,##0.00\\ "€"'
    FMT_PCT      = '0%'

    # Sheet1 — filas editadas en rojo, ocultar _editada
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    edit_col = headers.index("_editada") if "_editada" in headers else None
    for row in ws.iter_rows(min_row=2):
        if edit_col is not None and row[edit_col].value:
            for cell in row:
                cell.fill = ROJO_FONDO
                cell.font = ROJO_TEXTO
    if edit_col is not None:
        ws.column_dimensions[get_column_letter(edit_col + 1)].hidden = True

    # Formato € en columnas de precio (Sheet1, Fisio Invasiva)
    for ws_name in ["Sheet1", "Fisio Invasiva"]:
        if ws_name not in wb.sheetnames:
            continue
        ws2 = wb[ws_name]
        hdrs = [c.value for c in ws2[1]]
        for i, h in enumerate(hdrs, 1):
            if h in ("Precio", "Precio reserva", "Ingreso estimado", "Facturación"):
                col_l = get_column_letter(i)
                for cell in ws2[col_l][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = FMT_EURO

    # Ventas Pendientes — cabecera roja
    if "Ventas Pendientes" in wb.sheetnames:
        ws_vp = wb["Ventas Pendientes"]
        for cell in ws_vp[1]:
            cell.fill = CAB_ROJA
            cell.font = TEXTO_BLANCO
        hdrs_vp = [c.value for c in ws_vp[1]]
        for i, h in enumerate(hdrs_vp, 1):
            if h in ("Precio", "Precio reserva", "Precio (€)", "Precio reserva (€)"):
                col_l = get_column_letter(i)
                for cell in ws_vp[col_l][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = FMT_EURO

    # Sesiones — cabecera oscura + formato €/%
    if "Sesiones" in wb.sheetnames:
        ws_ses = wb["Sesiones"]
        for cell in ws_ses[1]:
            cell.fill = CAB_OSCURA
            cell.font = TEXTO_BLANCO
        hdrs_ses = [c.value for c in ws_ses[1]]
        for i, h in enumerate(hdrs_ses, 1):
            col_l = get_column_letter(i)
            if h in ("€/pers.", "€/ses."):
                for cell in ws_ses[col_l][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = FMT_EURO
            if h == "Ocupación":
                for cell in ws_ses[col_l][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = FMT_PCT
